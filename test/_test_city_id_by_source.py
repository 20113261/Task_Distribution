import pymongo
import openpyxl
from conf import config
from conn_pool import base_data_pool, source_info_pool, task_db_monitor_db_pool
from mysql_execute import fetchall, fetchall_ss
from collections import defaultdict

client = pymongo.MongoClient(host=config.mongo_host)
coll = client['RoutineBaseTask']['BaseTask_Hotel']

sql = '''select name from city where id={};'''
city_source_info = defaultdict(dict)
for package_id in range(5, 9):
    for line in coll.find({'package_id': package_id}):
        source = line['task_args']['source']
        suggest = line['task_args']['suggest']
        city_id = line['task_args']['city_id']
        if not city_source_info[city_id].get('city_name'):
            city_name = [ i for i in fetchall(source_info_pool, sql.format(city_id))][0][0]
        city_source_info[city_id][source] = suggest
        city_source_info[city_id]['city_name'] = city_name

for city_id, value in city_source_info.items():
    value['len'] = len(value.values()) - 1

excel = openpyxl.Workbook()
sheet = excel.create_sheet('1')
count = 1
sheet.cell(row=count, column=1, value='城市')
sheet.cell(row=count, column=2, value='城市id')
sheet.cell(row=count, column=3, value='ctrip')
sheet.cell(row=count, column=4, value='booking')
sheet.cell(row=count, column=5, value='expedia')
sheet.cell(row=count, column=6, value='hotels')
sheet.cell(row=count, column=7, value='elong')
sheet.cell(row=count, column=8, value='agoda')
sheet.cell(row=count, column=9, value='数量')


for city_id, value in city_source_info.items():
    count += 1
    print(count)
    # sheet.cell(row=count, column=1, value='4')

    sheet.cell(row=count, column=1, value=value['city_name'])
    sheet.cell(row=count, column=2, value=city_id)
    if value.get('ctripListHotel'):
        sheet.cell(row=count, column=3, value=1)
    if value.get('bookingListHotel'):
        sheet.cell(row=count, column=4, value=1)
    if value.get('expediaListHotel'):
        sheet.cell(row=count, column=5, value=1)
    if value.get('hotelsListHotel'):
        sheet.cell(row=count, column=6, value=1)
    if value.get('elongListHotel'):
        sheet.cell(row=count, column=7, value=1)
    if value.get('agodaListHotel'):
        sheet.cell(row=count, column=8, value=1)
    sheet.cell(row=count, column=9, value=value['len'])


excel.save('city_source_info.xlsx')
print(sheet.max_row)
print()